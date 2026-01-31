"""
gracetulsi_xlsx_pipeline.py

ETVL pipeline for SBA FY22 Home disaster loan data.
Computes total verified loss by state.
"""

from pathlib import Path
from typing import Any
import openpyxl


def _to_float(value: Any) -> float:
    """Convert Excel cell values to float safely."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def extract_state_verified_loss_rows(*, file_path: Path, sheet_name: str) -> list[tuple[str, float]]:
    """E: Extract (state_code, total_verified_loss) rows from an Excel sheet.

    Uses:
      H = state code (e.g., 'UT')
      I = total verified loss (numeric)

    Skips rows where state is missing or loss is missing/blank.
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Missing input file: {file_path}")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Found: {workbook.sheetnames}")

    sheet = workbook[sheet_name]

    rows: list[tuple[str, float]] = []

    # Data starts at row 6 per your sheet layout.
    for r in range(6, sheet.max_row + 1):
        state_val = sheet[f"H{r}"].value
        loss_val = sheet[f"I{r}"].value

        if state_val is None:
            continue

        state = str(state_val).strip()
        if not state:
            continue

        # Treat blank loss as skip (not zero), to avoid accidental misinterpretation.
        if loss_val is None or (isinstance(loss_val, str) and not loss_val.strip()):
            continue

        loss = _to_float(loss_val)
        rows.append((state, loss))

    return rows

def transform_total_verified_loss_by_state(
    *, rows: list[tuple[str, float]]
) -> dict[str, float]:
    """T: Sum total verified loss by state code."""
    totals: dict[str, float] = {}

    for state, loss in rows:
        # Add this row's loss into that state's running total
        totals[state] = totals.get(state, 0.0) + loss

    return totals

def verify_state_totals(*, totals: dict[str, float]) -> None:
    """V: Verify state total verified loss values are valid."""
    if not totals:
        raise ValueError("No state totals were computed.")

    for state, total in totals.items():
        if total < 0:
            raise ValueError(f"Negative total loss for state {state}: {total}")


def load_state_totals_report(
    *,
    totals: dict[str, float],
    out_path: Path,
    sheet_name: str,
    top_n: int = 10,
) -> None:
    """L: Write a text report summarizing total verified loss by state."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    ranked = sorted(totals.items(), key=lambda item: item[1], reverse=True)

    with out_path.open("w", encoding="utf-8") as f:
        f.write("SBA FY22 Home - Total Verified Loss by State\n")
        f.write("=" * 48 + "\n")
        f.write(f"Sheet: {sheet_name}\n")
        f.write(f"States counted: {len(totals)}\n")
        f.write(f"Top N: {top_n}\n\n")

        f.write("Rank | State | Total Verified Loss\n")
        f.write("-" * 40 + "\n")

        for i, (state, total) in enumerate(ranked[:top_n], start=1):
            f.write(f"{i:>4} | {state:<5} | {total:>18,.2f}\n")

def run_pipeline(*, raw_dir: Path, processed_dir: Path, logger: Any) -> None:
    """Run the full ETVL pipeline for SBA FY22 Home total verified loss by state."""
    logger.info("SBA XLSX: START")

    sheet = "FY22 Home"
    input_file = raw_dir / "sba_disaster_loan_data_fy22.xlsx"
    output_file = processed_dir / "gracetulsi_verified_loss_by_state.txt"

    rows = extract_state_verified_loss_rows(file_path=input_file, sheet_name=sheet)
    totals = transform_total_verified_loss_by_state(rows=rows)
    verify_state_totals(totals=totals)

    load_state_totals_report(
        totals=totals,
        out_path=output_file,
        sheet_name=sheet,
        top_n=10,
    )

    logger.info("SBA XLSX: wrote %s", output_file)
    logger.info("SBA XLSX: END")
